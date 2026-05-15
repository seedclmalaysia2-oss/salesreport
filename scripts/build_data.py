"""
Parse all xlsx files in `Sales & Forecast/` into one src/data.json.

Two file types:
  <SP> <YYYY> Sales Analysis by customer.xlsx
      Sheet 'Page 1'. Header at row 5: 'No' col 2, 'Customer Name' col 5,
      monthly Jan..Dec at cols 7,9,11,13,15,17,19,21,23,25,27,29 (odd).
      Data rows have integer 'No' in col 2; stop at 'Grand Total'.
  <SP> <YYYY> Stock Sales Analysis - Summary by [Bb]rand.xlsx
      Sheet 'Page 1'. Brand IDs in row 8 cols 14..(max-1). Last col = 'Total'.
      Each customer has an Amt row: col 2 = customer name, col 11 = 'Amt',
      cols 14..max-1 = revenue per brand. Stop at customer name 'Total'.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = ROOT / "Sales & Forecast"
OUT = ROOT / "src" / "data.json"

MONTH_COLS = [7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29]

FNAME_RE = re.compile(
    r"^(?P<sp>.+?) (?P<year>\d{4}) "
    r"(?P<kind>Sales Analysis by customer|Stock Sales Analysis - Summary by [Bb]rand)\.xlsx$"
)


def is_sales_brand(brand_id):
    """Brand IDs ending in FC (Free of Charge / boxes), T or TR (Trial Lens /
    pieces) carry no revenue and must not be counted as paid sales. Filter
    them out at parse time so they don't leak into any dashboard view.
    """
    if not brand_id or not isinstance(brand_id, str):
        return False
    bid = brand_id.strip().upper()
    if bid == "TOTAL" or not bid:
        return False
    if bid.endswith("FC"):
        return False
    if bid.endswith("TR"):
        return False
    if bid.endswith("T"):
        return False
    return True


def parse_customer_file(path: Path):
    """Returns list of {customer, months[12], total} dicts."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Page 1"]
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=False):
        no = row[1].value if len(row) > 1 else None  # col 2 (0-indexed col 1)
        if no is None or no == "":
            continue
        if isinstance(no, str) and "Total" in no:
            break
        # try to coerce to int
        try:
            int(no)
        except (TypeError, ValueError):
            continue
        name = row[4].value if len(row) > 4 else None  # col 5
        if not name:
            continue
        months = []
        for c in MONTH_COLS:
            v = row[c - 1].value if len(row) >= c else None
            months.append(float(v) if isinstance(v, (int, float)) else 0.0)
        total = sum(months)
        rows.append({"customer": str(name).strip(), "months": months, "total": round(total, 2)})
    wb.close()
    return rows


def parse_brand_file(path: Path):
    """Returns list of {customer, brand, amt} dicts.

    Column positions vary between files: 'AcStockBrandID' (and the first brand
    column under it) can sit at col 12, 13, or 14. The 'Amt' label can sit at
    col 9, 10, or 11. Detect both dynamically.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb["Page 1"]
    max_col = ws.max_column

    # Find first brand col: where row 7 cell == 'AcStockBrandID', brands run
    # from that col through max_col (last col is usually 'Total').
    first_brand_col = None
    for c in range(1, max_col + 1):
        if ws.cell(7, c).value == "AcStockBrandID":
            first_brand_col = c
            break
    if first_brand_col is None:
        # fallback: locate first non-empty brand-ID-looking cell on row 8
        for c in range(1, max_col + 1):
            v = ws.cell(8, c).value
            if isinstance(v, str) and v not in ("GroupingOption_ID", "MasterData") and v.strip():
                first_brand_col = c
                break
    if first_brand_col is None:
        return []

    brand_cols = {}
    for c in range(first_brand_col, max_col + 1):
        v = ws.cell(8, c).value
        if v and isinstance(v, str):
            brand_cols[c] = v.strip()

    # Find Amt label col: scan row 9..min(40, max_row) for cell value 'Amt'.
    amt_col = None
    for r in range(9, min(ws.max_row + 1, 50)):
        for c in range(1, first_brand_col):
            if ws.cell(r, c).value == "Amt":
                amt_col = c
                break
        if amt_col:
            break
    if amt_col is None:
        return []

    # Build a list of (customer, amt_row, qty_row) tuples by scanning sequentially.
    blocks = []
    pending_amt = None  # (customer, row_idx)
    for r in range(9, ws.max_row + 1):
        label = ws.cell(r, amt_col).value
        cust = ws.cell(r, 2).value
        if label == "Amt":
            if cust:
                cust_str = str(cust).strip()
                if cust_str.lower() == "total":
                    break
                pending_amt = (cust_str, r)
        elif label == "Qty" and pending_amt is not None:
            blocks.append((pending_amt[0], pending_amt[1], r))
            pending_amt = None
    if pending_amt is not None:
        # Customer with Amt row but no matching Qty row in sheet
        blocks.append((pending_amt[0], pending_amt[1], None))

    out = []
    skipped_brands = set()
    for customer, amt_row, qty_row in blocks:
        for c, brand in brand_cols.items():
            if not is_sales_brand(brand):
                skipped_brands.add(brand)
                continue
            amt_v = ws.cell(amt_row, c).value
            qty_v = ws.cell(qty_row, c).value if qty_row else None
            amt = float(amt_v) if isinstance(amt_v, (int, float)) and amt_v != 0 else 0.0
            qty = float(qty_v) if isinstance(qty_v, (int, float)) and qty_v != 0 else 0.0
            if amt != 0 or qty != 0:
                out.append({"customer": customer, "brand": brand, "amt": amt, "qty": qty})
    wb.close()
    return out, skipped_brands


def main():
    if not SRC_DIR.exists():
        print(f"ERROR: {SRC_DIR} not found", file=sys.stderr)
        sys.exit(1)

    customers_data = []  # list of {sp, year, customer, months[12], total}
    brand_data = []  # list of {sp, year, customer, brand, amt}

    files = sorted(SRC_DIR.glob("*.xlsx"))
    print(f"Found {len(files)} xlsx files")

    customer_count = 0
    brand_count = 0
    all_skipped_brands = set()

    for f in files:
        m = FNAME_RE.match(f.name)
        if not m:
            print(f"  SKIP (unrecognized name): {f.name}")
            continue
        sp = m.group("sp").strip()
        year = int(m.group("year"))
        kind = m.group("kind")

        try:
            if kind == "Sales Analysis by customer":
                rows = parse_customer_file(f)
                for r in rows:
                    customers_data.append({"sp": sp, "year": year, **r})
                customer_count += 1
                print(f"  customer: {f.name} -> {len(rows)} rows")
            else:  # brand file
                rows, skipped = parse_brand_file(f)
                for r in rows:
                    brand_data.append({"sp": sp, "year": year, **r})
                brand_count += 1
                all_skipped_brands.update(skipped)
                print(f"  brand:    {f.name} -> {len(rows)} rows (skipped {len(skipped)} non-sales codes)")
        except Exception as e:
            print(f"  ERROR parsing {f.name}: {e}", file=sys.stderr)
            raise

    # Derive SP/year totals from customers_data
    summary_map = {}
    for r in customers_data:
        key = (r["sp"], r["year"])
        if key not in summary_map:
            summary_map[key] = {
                "sp": r["sp"],
                "year": r["year"],
                "total": 0.0,
                "customers": 0,
                "months": [0.0] * 12,
            }
        s = summary_map[key]
        s["total"] += r["total"]
        if r["total"] > 0:
            s["customers"] += 1
        for i, m in enumerate(r["months"]):
            s["months"][i] += m

    summary = []
    for s in summary_map.values():
        s["total"] = round(s["total"], 2)
        s["months"] = [round(m, 2) for m in s["months"]]
        summary.append(s)
    summary.sort(key=lambda s: (s["sp"], s["year"]))

    # Top customers all-time
    cust_totals = {}
    for r in customers_data:
        cust_totals[r["customer"]] = cust_totals.get(r["customer"], 0.0) + r["total"]
    top_customers = sorted(
        [{"customer": c, "total": round(t, 2)} for c, t in cust_totals.items()],
        key=lambda x: -x["total"],
    )[:20]

    # Unique salespeople, years, brands
    salespeople = sorted({r["sp"] for r in customers_data})
    years = sorted({r["year"] for r in customers_data})
    brands = sorted({r["brand"] for r in brand_data})

    payload = {
        "salespeople": salespeople,
        "years": years,
        "brands": brands,
        "summary": summary,
        "topCustomers": top_customers,
        "customers": customers_data,  # raw per-customer monthly
        "brandSales": brand_data,  # raw per-customer per-brand annual
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, separators=(",", ":"), ensure_ascii=False)

    size_kb = OUT.stat().st_size / 1024
    print()
    print(f"Wrote {OUT.relative_to(ROOT)} ({size_kb:.1f} KB)")
    print(f"  customer files parsed: {customer_count}")
    print(f"  brand files parsed:    {brand_count}")
    print(f"  customers rows:        {len(customers_data)}")
    print(f"  brand sales rows:      {len(brand_data)}")
    print(f"  unique salespeople:    {len(salespeople)} ({salespeople})")
    print(f"  years:                 {years}")
    print(f"  unique brands:         {len(brands)} (sales-only — FC/T/TR excluded)")
    print(f"  non-sales brand codes: {len(all_skipped_brands)} excluded ({sorted(all_skipped_brands)[:8]}{'...' if len(all_skipped_brands) > 8 else ''})")
    print(f"  summary rows (sp×yr):  {len(summary)}")


if __name__ == "__main__":
    main()
