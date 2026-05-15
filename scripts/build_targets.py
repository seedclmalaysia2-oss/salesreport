"""
Parse Sales Target xlsx files and emit per-rep monthly targets to src/targets.json.

Files of interest (all in 'Sales Target/'):
  SEED(M) Sales Summary By Salesman <Mon>_<YYYY>.xlsx

Layout in those files (row indices 1-based):
  r4: month headers in cols 2..13 = JAN..DEC
  r5: 'Jpn Sales Target' values in cols 2..13 → team monthly target
  r11: 'Salesman' header row, with rep names in cols 2..N
  r12..r17 / r19..r24: monthly per-rep actuals (we re-derive contribution % from these)

We extract:
  • team_targets[year][month_idx 0..11] = team total target
  • rep_actuals[year][rep_name][month_idx] = actual sales (used for contribution %)

Then derive per-rep targets via prior-year contribution %:
  target[year, rep, month] = team_target[year, month] * contribution_pct[year-1, rep]
For the earliest year that has no prior year, fall back to that year's own contribution.

Outputs src/targets.json with shape:
  { team: [{year, month, target}], reps: [{year, month, sp, target}] }
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "Sales Target"
OUT = ROOT / "src" / "targets.json"

# Map "Alan Loh" → "Alan" (the SP name we use in the dashboard).
NAME_MAP = {
    "Alan Loh": "Alan",
    "Dino Lim": "Dino",
    "Khen Tan": "Khen",
    "Sakinah": "Sakinah",
    "Sakinah ": "Sakinah",
    "Simon Low": "Simon",
}


def find_year_from_title(title: str) -> int | None:
    m = re.search(r"(\d{4})\b", title or "")
    return int(m.group(1)) if m else None


def parse_salesman_file(path: Path):
    """Returns (year, team_monthly_targets[12], rep_actuals dict)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "Monthly Sales" in s), wb.sheetnames[0])
    ws = wb[sheet_name]

    title = ws.cell(1, 1).value or ""
    year = find_year_from_title(title)

    # Row 5 has the team monthly targets in cols 2..13
    targets = []
    for c in range(2, 14):
        v = ws.cell(5, c).value
        targets.append(float(v) if isinstance(v, (int, float)) else 0.0)

    # Row 11 cols 2..N has rep names; find which columns belong to mapped reps.
    rep_cols = {}
    for c in range(2, ws.max_column + 1):
        name = ws.cell(11, c).value
        if isinstance(name, str) and name.strip() in NAME_MAP:
            rep_cols[NAME_MAP[name.strip()]] = c

    # Rep monthly actuals: rows 12..17 (Jan..Jun), 19..24 (Jul..Dec).
    rep_actuals = {sp: [0.0] * 12 for sp in rep_cols}
    month_rows = list(range(12, 18)) + list(range(19, 25))  # 12 rows, one per month
    for month_idx, r in enumerate(month_rows):
        for sp, c in rep_cols.items():
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                rep_actuals[sp][month_idx] = float(v)

    wb.close()
    return year, targets, rep_actuals


def main() -> int:
    files = sorted(SRC.glob("SEED(M) Sales Summary By Salesman*.xlsx"))
    if not files:
        print(f"No salesman summary files found in {SRC}", file=sys.stderr)
        return 1

    # Each year may have multiple snapshots (Mar_2026 + Apr_2026); the latest wins.
    by_year = {}
    for f in files:
        try:
            year, team, reps = parse_salesman_file(f)
        except Exception as e:
            print(f"  skip {f.name}: {e}")
            continue
        if not year:
            continue
        # Track most recent file mtime to pick the latest snapshot.
        mtime = f.stat().st_mtime
        prev = by_year.get(year)
        if prev is None or prev["mtime"] < mtime:
            by_year[year] = {"team": team, "reps": reps, "mtime": mtime, "src": f.name}
        print(f"  parsed {f.name}: year={year} team_target_jan={team[0]:.0f}")

    print(f"\nYears covered: {sorted(by_year)}")

    # Compute contribution % per year per rep (from actuals)
    contribution = {}  # year -> {sp: pct}
    for year, data in by_year.items():
        total = sum(sum(months) for months in data["reps"].values())
        if total <= 0:
            contribution[year] = {sp: 0 for sp in data["reps"]}
            continue
        contribution[year] = {
            sp: sum(months) / total for sp, months in data["reps"].items()
        }

    # Derive per-rep monthly targets using prior-year contribution.
    # If no prior year, use this year's own contribution as fallback.
    team_rows = []
    rep_rows = []
    sorted_years = sorted(by_year)
    for year in sorted_years:
        team_targets = by_year[year]["team"]
        # Team-level rows
        for m, t in enumerate(team_targets, start=1):
            if t > 0:
                team_rows.append({"year": year, "month": m, "sp": "_TEAM", "target_amt": round(t, 2)})

        # Find a basis for the contribution split
        prior_year = year - 1
        basis = contribution.get(prior_year) or contribution.get(year) or {}
        # Re-normalise the basis to only include SPs that have at least 1% (filter noise like Wani/Jerry)
        basis = {sp: pct for sp, pct in basis.items() if pct >= 0.01}
        if basis:
            s = sum(basis.values())
            basis = {sp: pct / s for sp, pct in basis.items()}

        if not basis:
            print(f"  ! no contribution basis for {year} — skipping per-rep derivation")
            continue

        for sp, pct in basis.items():
            for m, t in enumerate(team_targets, start=1):
                if t > 0:
                    rep_rows.append({
                        "year": year,
                        "month": m,
                        "sp": sp,
                        "target_amt": round(t * pct, 2),
                    })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump({"team": team_rows, "reps": rep_rows}, fh, separators=(",", ":"))

    # Print summary
    print(f"\nWrote {OUT.relative_to(ROOT)}")
    print(f"  team rows:    {len(team_rows)}")
    print(f"  per-rep rows: {len(rep_rows)}")
    for year in sorted_years:
        ann_team = sum(by_year[year]["team"])
        ann_per_rep = {}
        for r in rep_rows:
            if r["year"] == year:
                ann_per_rep[r["sp"]] = ann_per_rep.get(r["sp"], 0) + r["target_amt"]
        rep_str = ", ".join(f"{sp}={v/1000:.0f}K" for sp, v in sorted(ann_per_rep.items()))
        print(f"  {year}: team {ann_team/1_000_000:.2f}M  →  {rep_str}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
